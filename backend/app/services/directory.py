"""Team directory + identity resolution.

A workspace's directory is the roster of people that meeting decisions can be
assigned to and chased. A directory member is a `User` row (with
`is_login_enabled=False` when they only exist as a contact, not a dashboard
login) linked to a workspace via `WorkspaceMember`, plus optional
`ExternalIdentity` rows that map them to their Slack / Jira / GitHub / Linear
accounts so we can assign tickets and send notifications to the right person.

This module is the single place that turns a loose name spoken in a meeting
(e.g. "Dave") into a concrete directory member, and that creates/updates those
members from manual input, CSV/XLSX import, or a Slack directory sync.
"""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from ..models.core import (
    ExternalIdentity,
    MemberRole,
    User,
    WorkspaceMember,
)

# Providers we can store an external account id for, and the column aliases we
# accept for each when importing a spreadsheet.
PROVIDER_COLUMNS = {
    "slack": ("slack", "slack_id", "slack_user_id", "slack member id"),
    "jira": ("jira", "jira_account_id", "jira accountid", "jira id"),
    "github": ("github", "github_login", "github username", "gh"),
    "linear": ("linear", "linear_id", "linear user id"),
}


class DirectoryError(RuntimeError):
    pass


@dataclass
class MemberInput:
    name: str
    email: str | None = None
    title: str | None = None
    department: str | None = None
    role: str = "member"
    skills: list[str] = field(default_factory=list)
    aliases: list[str] = field(default_factory=list)
    external_ids: dict[str, str] = field(default_factory=dict)


def _norm(value: str | None) -> str:
    return (value or "").strip().casefold()


def _email_local(email: str | None) -> str:
    return _norm(email).split("@", 1)[0] if email else ""


async def resolve_owner(
    session: AsyncSession, workspace_id, owner_name: str | None
) -> User | None:
    """Map a free-text owner name from a transcript to a directory member.

    Matching order: exact display name, alias, then email local-part. All
    comparisons are case-insensitive. Returns None when there's no confident
    match (callers keep the task unassigned rather than guessing)."""
    if not owner_name or not owner_name.strip():
        return None
    members = (
        (
            await session.execute(
                select(User)
                .join(WorkspaceMember, WorkspaceMember.user_id == User.id)
                .where(WorkspaceMember.workspace_id == workspace_id)
            )
        )
        .scalars()
        .all()
    )
    target = _norm(owner_name)
    for user in members:
        if _norm(user.display_name) == target:
            return user
    for user in members:
        if target in {_norm(a) for a in (user.aliases or [])}:
            return user
    for user in members:
        if _email_local(user.email) == target:
            return user
    return None


async def upsert_member(
    session: AsyncSession, workspace_id, data: MemberInput
) -> User:
    """Create or update a directory member (matched by email, else by name)."""
    if not data.name.strip():
        raise DirectoryError("Member name is required")
    user: User | None = None
    if data.email:
        user = (
            await session.execute(select(User).where(User.email == data.email.strip()))
        ).scalar_one_or_none()
    if not user:
        # Match an existing directory-only contact in this workspace by name.
        user = await resolve_owner(session, workspace_id, data.name)
    if not user:
        user = User(
            email=(data.email or _placeholder_email(data.name, workspace_id)).strip(),
            display_name=data.name.strip(),
            is_login_enabled=False,
        )
        session.add(user)
        await session.flush()
    user.display_name = data.name.strip()
    if data.email:
        user.email = data.email.strip()
    if data.title is not None:
        user.title = data.title
    if data.department is not None:
        user.department = data.department
    if data.skills:
        user.skills = data.skills
    if data.aliases:
        user.aliases = sorted({*(user.aliases or []), *data.aliases})

    membership = (
        await session.execute(
            select(WorkspaceMember).where(
                WorkspaceMember.workspace_id == workspace_id,
                WorkspaceMember.user_id == user.id,
            )
        )
    ).scalar_one_or_none()
    role = data.role if data.role in {r.value for r in MemberRole} else "member"
    if not membership:
        session.add(
            WorkspaceMember(
                workspace_id=workspace_id, user_id=user.id, role=MemberRole(role)
            )
        )
    else:
        membership.role = MemberRole(role)

    for provider, external_id in data.external_ids.items():
        if external_id and external_id.strip():
            await _link_identity(session, user, provider, external_id.strip())
    await session.flush()
    return user


def _placeholder_email(name: str, workspace_id) -> str:
    stem = re.sub(r"[^a-z0-9]+", ".", name.lower()).strip(".") or "member"
    return f"{stem}.{str(workspace_id)[:8]}@directory.local"


async def _link_identity(
    session: AsyncSession, user: User, provider: str, external_user_id: str
) -> None:
    existing = (
        await session.execute(
            select(ExternalIdentity).where(
                ExternalIdentity.user_id == user.id,
                ExternalIdentity.provider == provider,
            )
        )
    ).scalar_one_or_none()
    if existing:
        existing.external_user_id = external_user_id
    else:
        session.add(
            ExternalIdentity(
                user_id=user.id,
                provider=provider,
                external_user_id=external_user_id,
            )
        )


async def external_id_for(
    session: AsyncSession, user_id, provider: str
) -> str | None:
    identity = (
        await session.execute(
            select(ExternalIdentity).where(
                ExternalIdentity.user_id == user_id,
                ExternalIdentity.provider == provider,
            )
        )
    ).scalar_one_or_none()
    return identity.external_user_id if identity else None


# ---------------------------------------------------------------------------
# Spreadsheet parsing (CSV + XLSX)
# ---------------------------------------------------------------------------

IMPORT_TEMPLATE_HEADERS = [
    "name",
    "email",
    "title",
    "department",
    "role",
    "skills",
    "aliases",
    "slack_id",
    "jira_account_id",
    "github_login",
    "linear_id",
]


def _rows_to_members(rows: list[dict[str, str]]) -> list[MemberInput]:
    members: list[MemberInput] = []
    for raw in rows:
        row = {(_norm(k)): (v or "").strip() for k, v in raw.items() if k}
        name = row.get("name") or row.get("full name") or row.get("display_name")
        if not name:
            continue
        external_ids = {}
        for provider, columns in PROVIDER_COLUMNS.items():
            for column in columns:
                if row.get(_norm(column)):
                    external_ids[provider] = row[_norm(column)]
                    break
        members.append(
            MemberInput(
                name=name,
                email=row.get("email") or None,
                title=row.get("title") or None,
                department=row.get("department") or row.get("team") or None,
                role=(row.get("role") or "member").lower(),
                skills=_split_list(row.get("skills")),
                aliases=_split_list(row.get("aliases")),
                external_ids=external_ids,
            )
        )
    return members


def _split_list(value: str | None) -> list[str]:
    if not value:
        return []
    return [part.strip() for part in re.split(r"[;,|]", value) if part.strip()]


def parse_members_file(filename: str, content: bytes) -> list[MemberInput]:
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return _rows_to_members(_parse_xlsx(content))
    if name.endswith(".csv") or not name:
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        return _rows_to_members(list(reader))
    raise DirectoryError("Unsupported file type; upload a .csv or .xlsx file")


def _parse_xlsx(content: bytes) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:  # pragma: no cover
        raise DirectoryError(
            "Excel import requires the 'openpyxl' package"
        ) from error
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header = [str(cell).strip() if cell is not None else "" for cell in next(rows_iter)]
    except StopIteration:
        return []
    rows: list[dict[str, str]] = []
    for values in rows_iter:
        if values is None or all(cell is None for cell in values):
            continue
        row = {
            header[i]: ("" if value is None else str(value))
            for i, value in enumerate(values)
            if i < len(header) and header[i]
        }
        rows.append(row)
    return rows
